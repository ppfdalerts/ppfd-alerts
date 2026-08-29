"""Import the daily apparatus roster text export into roster_units JSON files."""

import argparse
import datetime as dt
from io import BytesIO
import json
import re
from zipfile import ZipFile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


DATE_RE = re.compile(r"^=====\s+(\d{2})/(\d{2})/(\d{4})\s+=====$")
# TeleStaff text exports vary: some include a station note before the slash
# and some omit the trailing colon entirely.
UNIT_RE = re.compile(
    r"^Station\s+\d+(?:\s+\{.*?\})?\s*/\s*([^:{]+?)(?:\s+\{.*)?\s*:?\s*$"
)
FIELD_RE = re.compile(r"^\s*(.+?)\s+-\s+(.+?)\s*$")
CLOCK_RE = re.compile(r"^\d{1,2}:\d{2}$")
NUMBER_RE = re.compile(r"^\d+(?:\.\d+)?$")

UNIT_CODES = {
    "Truck 33": "T33",
    "Rescue 33": "R33",
    "LR36": "LR36",
    "Engine 34": "E34",
    "Rescue 34": "R34",
    "District 35": "D35",
    "Rescue 35": "R35",
    "Squad 36": "S36",
}

XLSX_APPARATUS_CODES = {
    "Truck 33": "T33",
    "Rescue 33": "R33",
    "LR36": "LR36",
    "Engine 34": "E34",
    "Rescue 34": "R34",
    "District 35": "D35",
    "Engine 35": "E35",
    "Engine36": "E36",
    "Engine 36": "E36",
    "Rescue 35": "R35",
    "Squad 36": "S36",
}

WATCH_UNIT_CODES = {
    "D35", "E34", "E35", "E36", "LR36", "R33", "R34", "R35", "S36", "T33"
}


def canonical_name(value: str) -> str:
    """Remove certifications so a person's display name remains stable by date."""
    name = value.split("(", 1)[0].strip()
    return re.sub(r"\s+", " ", name)


def name_signature(value: str) -> tuple[str, str]:
    base = canonical_name(value).lower()
    if "," not in base:
        return base, ""
    last, first = base.split(",", 1)
    return last.strip(), first.strip().split()[0] if first.strip() else ""


def load_existing_people(roster_dir: Path) -> dict[str, str]:
    """Map text names to known personnel IDs using existing roster JSON files."""
    exact: dict[str, str] = {}
    signatures: dict[tuple[str, str], set[str]] = {}
    for path in sorted(roster_dir.glob("roster_units_*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        for unit in payload.get("units", []) or []:
            for entry in unit.get("entries", []) or []:
                pid = str(entry.get("id") or "").strip()
                name = canonical_name(str(entry.get("name") or ""))
                if not pid or not name:
                    continue
                exact[name.lower()] = pid
                last, first = name_signature(name)
                signatures.setdefault((last, first), set()).add(pid)

    result: dict[str, str] = {}
    for name in exact:
        result[name] = exact[name]
    return result, signatures


def resolve_person_id(name: str, known: tuple[dict[str, str], dict[tuple[str, str], set[str]]]) -> str:
    exact, signatures = known
    canonical = canonical_name(name)
    exact_id = exact.get(canonical.lower())
    if exact_id:
        return exact_id
    last, first = name_signature(canonical)
    candidates: set[str] = set()
    for (candidate_last, candidate_first), ids in signatures.items():
        if candidate_last != last:
            continue
        if first and candidate_first and (
            candidate_first.startswith(first) or first.startswith(candidate_first)
        ):
            candidates.update(ids)
    if len(candidates) == 1:
        return next(iter(candidates))
    return canonical


def _duration_hours(start: str, end: str) -> float:
    sh, sm = (int(part) for part in start.split(":", 1))
    eh, em = (int(part) for part in end.split(":", 1))
    start_min = sh * 60 + sm
    end_min = eh * 60 + em
    if end_min <= start_min:
        end_min += 24 * 60
    return round((end_min - start_min) / 60.0, 2)


def _add_hours(start: str, hours: float) -> str:
    sh, sm = (int(part) for part in start.split(":", 1))
    end_min = int(round(sh * 60 + sm + hours * 60)) % (24 * 60)
    return f"{end_min // 60:02d}:{end_min % 60:02d}"


def parse_roster(path: Path, known_people):
    by_date: dict[dt.date, dict[str, dict]] = {}
    current_date: dt.date | None = None
    current_unit: dict | None = None
    current_entries: list[dict] = []
    open_entry: dict | None = None
    next_start = "07:30"
    transition_pending = False

    def close_entry(end: str, explicit_hours: float | None = None) -> None:
        nonlocal open_entry, next_start
        if open_entry is None:
            next_start = end
            return
        open_entry["through"] = end
        open_entry["hours"] = (
            round(explicit_hours, 2)
            if explicit_hours is not None
            else _duration_hours(open_entry["from"], end)
        )
        current_entries.append(open_entry)
        open_entry = None
        next_start = end

    def finish_unit() -> None:
        nonlocal current_unit, current_entries, open_entry, next_start, transition_pending
        if current_unit is None or current_date is None:
            return
        close_entry("07:30")
        unique_entries = []
        seen_intervals = set()
        for entry in current_entries:
            key = (
                str(entry.get("id") or ""),
                str(entry.get("from") or ""),
                str(entry.get("through") or ""),
            )
            if key in seen_intervals:
                continue
            seen_intervals.add(key)
            unique_entries.append(entry)
        current_entries = unique_entries
        current_unit["entries"] = current_entries
        by_date[current_date][current_unit["unit_code"]] = current_unit
        current_unit = None
        current_entries = []
        open_entry = None
        next_start = "07:30"
        transition_pending = False

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.rstrip()
        date_match = DATE_RE.match(line)
        if date_match:
            finish_unit()
            current_date = dt.date(
                int(date_match.group(3)), int(date_match.group(1)), int(date_match.group(2))
            )
            by_date.setdefault(current_date, {})
            continue

        unit_match = UNIT_RE.match(line)
        if unit_match:
            finish_unit()
            unit_name = unit_match.group(1).strip()
            unit_code = UNIT_CODES.get(unit_name)
            if unit_code and current_date is not None:
                current_unit = {
                    "shift": "Fire Department / Suppression / Imported apparatus roster",
                    "unit_name": unit_name,
                    "unit_code": unit_code,
                    "entries": [],
                }
                current_entries = []
                open_entry = None
                next_start = "07:30"
                transition_pending = False
            continue

        if current_unit is None:
            continue
        field_match = FIELD_RE.match(line)
        if not field_match:
            continue
        left = field_match.group(1).strip()
        right = field_match.group(2).strip()

        # A clock on the left ends the current interval. A numeric value on
        # the right is the ID for a same-position replacement; other values
        # start the next apparatus position.
        if CLOCK_RE.fullmatch(left):
            close_entry(left)
            transition_pending = bool(re.fullmatch(r"\d+|\?,\?", right))
            continue

        if "," not in right:
            continue

        person_name = canonical_name(right)

        # Numeric left values are the completed duration of the prior person,
        # while the name on the right is the replacement in the same position.
        if NUMBER_RE.fullmatch(left):
            hours = float(left)
            if not transition_pending and open_entry is not None:
                end = _add_hours(open_entry["from"], hours)
                close_entry(end, explicit_hours=hours)
            open_entry = {
                "rank": "",
                "id": resolve_person_id(person_name, known_people),
                "name": person_name,
                "code": "",
                "from": next_start,
            }
            transition_pending = False
            continue

        # A normal rank/name row starts a new parallel apparatus position. If
        # another replacement was still open, it runs through shift end.
        if open_entry is not None:
            close_entry("07:30")
        if person_name:
            person_name = canonical_name(right)
            open_entry = {
                "rank": left,
                "id": resolve_person_id(person_name, known_people),
                "name": person_name,
                "code": "",
                "from": next_start if transition_pending else "07:30",
            }
            next_start = open_entry["from"]
            transition_pending = False

    finish_unit()
    parsed_entries = sum(
        len(unit.get("entries") or [])
        for date_units in by_date.values()
        for unit in date_units.values()
    )
    # The columnar export should produce many entries per date. A few
    # accidental matches from the legacy parser are not sufficient to treat
    # that format as successfully parsed.
    if parsed_entries < len(by_date):
        return _parse_roster_columnar(path, known_people)
    return by_date


def _parse_roster_columnar(path: Path, known_people):
    """Parse the line-per-field format used by current TeleStaff exports."""
    by_date: dict[dt.date, dict[str, dict]] = {}
    current_date: dt.date | None = None
    current_unit: dict | None = None
    history: list[str] = []
    lines = path.read_text(encoding="utf-8-sig").splitlines()

    def finish_unit() -> None:
        nonlocal current_unit, history
        if current_unit is not None and current_date is not None:
            by_date[current_date][current_unit["unit_code"]] = current_unit
        current_unit = None
        history = []

    def is_numeric(value: str) -> bool:
        return bool(re.fullmatch(r"\d+", value))

    def is_clock(value: str) -> bool:
        return bool(re.fullmatch(r"\d{1,2}:\d{2}", value))

    def is_rank(value: str) -> bool:
        return bool(value) and not is_numeric(value) and not is_clock(value) and value not in {
            "Rank", "ID", "Name", "From", "Through"
        }

    for index, raw_line in enumerate(lines):
        line = raw_line.strip()
        date_match = DATE_RE.match(line)
        if date_match:
            finish_unit()
            current_date = dt.date(
                int(date_match.group(3)), int(date_match.group(1)), int(date_match.group(2))
            )
            by_date.setdefault(current_date, {})
            continue

        unit_match = UNIT_RE.match(line)
        if unit_match:
            finish_unit()
            unit_name = unit_match.group(1).strip()
            unit_code = UNIT_CODES.get(unit_name)
            if unit_code and current_date is not None:
                current_unit = {
                    "shift": "Fire Department / Suppression / Imported apparatus roster",
                    "unit_name": unit_name,
                    "unit_code": unit_code,
                    "entries": [],
                }
            continue

        if current_unit is None or not line:
            continue

        # Names are the only roster fields containing a comma. The next four
        # non-empty fields are code, start, end, and scheduled hours.
        if "," not in line or line in {"-,-", "?,?"}:
            if line:
                history.append(line)
            continue

        following: list[str] = []
        for candidate in lines[index + 1 :]:
            candidate = candidate.strip()
            if not candidate:
                continue
            if DATE_RE.match(candidate) or UNIT_RE.match(candidate):
                break
            following.append(candidate)
            if len(following) == 4:
                break
        if len(following) < 3:
            history.append(line)
            continue

        person_name = canonical_name(line)
        person_id = ""
        rank = ""
        if history and (is_numeric(history[-1]) or history[-1] in {"-,-", "?,?"}):
            person_id = history[-1]
            if len(history) >= 2 and is_rank(history[-2]):
                rank = history[-2]
        entry = {
            "rank": rank,
            "id": resolve_person_id(person_name, known_people) if person_id in {"", "-,-", "?,?"} else person_id,
            "name": person_name,
            "code": following[0],
            "from": _time_text(following[1]),
            "through": _time_text(following[2]),
            "hours": float(following[3]) if len(following) >= 4 and NUMBER_RE.fullmatch(following[3]) else None,
        }
        current_unit["entries"].append(entry)
        history.append(line)

    finish_unit()
    return by_date


def _time_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        value = value.time()
    if isinstance(value, dt.time):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = str(value).strip()
    match = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?", text)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return text


def parse_apparatus_assignments_xlsx(sheet, known_people):
    """Parse the Kronos seven-day Apparatus Assignments worksheet."""
    by_date: dict[dt.date, dict[str, dict]] = {}
    seen: set[tuple] = set()
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return by_date
    columns = {
        str(value or "").strip().lower(): index
        for index, value in enumerate(header)
    }

    def value(row, *names):
        for name in names:
            index = columns.get(name.lower())
            if index is not None and index < len(row):
                return row[index]
        return None

    for row in rows:
        date_value = value(row, "Date")
        apparatus = value(row, "Apparatus")
        position = value(row, "Position")
        name = value(row, "Personnel")
        employee_id = value(row, "Employee ID", "Personnel ID")
        code = value(row, "Work Code")
        start = value(row, "From")
        through = value(row, "Through")
        hours = value(row, "Duration")
        unit_code = XLSX_APPARATUS_CODES.get(str(apparatus or "").strip())
        if not unit_code or not name or str(name).strip() in ("-,-", "?,?"):
            continue
        if isinstance(date_value, dt.datetime):
            roster_date = date_value.date()
        elif isinstance(date_value, dt.date):
            roster_date = date_value
        else:
            try:
                roster_date = dt.date.fromisoformat(str(date_value).strip()[:10])
            except ValueError:
                continue

        person_name = str(name).strip()
        canonical = canonical_name(person_name)
        supplied_id = str(employee_id or "").strip()
        person_id = supplied_id if supplied_id.isdigit() else resolve_person_id(canonical, known_people)
        rank = str(position or "").strip()
        from_time = _time_text(start)
        through_time = _time_text(through)
        if hours not in (None, ""):
            duration = float(hours)
        elif start and through:
            duration = _duration_hours(from_time, through_time)
        else:
            duration = None
        entry_key = (roster_date, unit_code, person_id, rank, str(code or ""), from_time, through_time, duration)
        if entry_key in seen:
            continue
        seen.add(entry_key)

        date_units = by_date.setdefault(roster_date, {})
        unit = date_units.setdefault(
            unit_code,
            {
                "shift": "Fire Department / Suppression / XLSX apparatus roster",
                "unit_name": next((name for name, code_value in UNIT_CODES.items() if code_value == unit_code), unit_code),
                "unit_code": unit_code,
                "entries": [],
            },
        )
        unit["entries"].append(
            {
                "rank": rank,
                "id": person_id,
                "name": canonical,
                "code": str(code or "").strip(),
                "from": from_time,
                "through": through_time,
                "hours": duration,
            }
        )
    return by_date


def _xlsx_bytes(path: Path) -> tuple[str, bytes]:
    if path.suffix.lower() != ".zip":
        return path.name, path.read_bytes()
    with ZipFile(path) as archive:
        candidates = [
            info for info in archive.infolist()
            if info.filename.lower().endswith((".xlsx", ".xlsm"))
        ]
        if not candidates:
            raise SystemExit("ZIP does not contain an XLSX roster document.")
        info = max(candidates, key=lambda item: item.file_size)
        return Path(info.filename).name, archive.read(info)


def parse_roster_xlsx(path: Path, known_people):
    if load_workbook is None:
        raise SystemExit("openpyxl is required to import the XLSX roster.")
    source_name, data = _xlsx_bytes(path)
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    if "Daily Roster" not in workbook.sheetnames:
        if "Apparatus Assignments" in workbook.sheetnames:
            return parse_apparatus_assignments_xlsx(workbook["Apparatus Assignments"], known_people), source_name
        raise SystemExit("XLSX does not contain a Daily Roster or Apparatus Assignments sheet.")
    sheet = workbook["Daily Roster"]
    by_date: dict[dt.date, dict[str, dict]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row) + [None] * 10
        date_value, apparatus, unit_code, rank, name, code, start, through, hours = values[:9]
        if not date_value or str(unit_code or "").strip().upper() not in WATCH_UNIT_CODES:
            continue
        if not name or str(name).strip() in ("-,-", "?,?"):
            continue
        if isinstance(date_value, dt.datetime):
            roster_date = date_value.date()
        elif isinstance(date_value, dt.date):
            roster_date = date_value
        else:
            roster_date = dt.date.fromisoformat(str(date_value).strip()[:10])
        code = str(unit_code).strip().upper()
        unit_name = str(apparatus or code).split("/", 1)[-1].strip() or code
        person_name = str(name).strip()
        canonical = canonical_name(person_name)
        entry = {
            "rank": str(rank or "").strip(),
            "id": resolve_person_id(canonical, known_people),
            "name": person_name,
            "code": str(code or "").strip(),
            "from": _time_text(start),
            "through": _time_text(through),
            "hours": float(hours) if hours not in (None, "") else None,
        }
        date_units = by_date.setdefault(roster_date, {})
        unit = date_units.setdefault(
            code,
            {
                "shift": "Fire Department / Suppression / XLSX apparatus roster",
                "unit_name": unit_name,
                "unit_code": code,
                "entries": [],
            },
        )
        unit["entries"].append(entry)
    return by_date, source_name


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="apparatus-roster TXT export")
    parser.add_argument("--out-dir", type=Path, default=Path("TSlogs"))
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--overwrite-imported",
        action="store_true",
        help="Replace only roster files previously generated from this same input file.",
    )
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"Roster text file not found: {args.input}")
    args.out_dir.mkdir(parents=True, exist_ok=True)
    known_people = load_existing_people(args.out_dir)
    if args.input.suffix.lower() in (".zip", ".xlsx", ".xlsm"):
        by_date, source_name = parse_roster_xlsx(args.input, known_people)
    else:
        by_date = parse_roster(args.input, known_people)
        source_name = args.input.name

    written = 0
    skipped = 0
    for date in sorted(by_date):
        output = args.out_dir / f"roster_units_{date:%Y-%m-%d}.json"
        if output.exists():
            existing_source = ""
            try:
                existing_source = str(json.loads(output.read_text(encoding="utf-8")).get("source_file") or "")
            except Exception:
                pass
            replace_existing = bool(args.overwrite)
            if args.overwrite_imported:
                replace_existing = existing_source.startswith("apparatus-roster-")
            if not replace_existing:
                skipped += 1
                continue
        units = [by_date[date][code] for code in sorted(by_date[date])]
        payload = {
            "date": date.isoformat(),
            "source_file": source_name,
            "units": units,
        }
        tmp = output.with_suffix(output.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
        tmp.replace(output)
        written += 1

    print(f"Roster dates parsed: {len(by_date)}")
    print(f"Roster files written: {written}")
    print(f"Existing roster files skipped: {skipped}")
    if by_date:
        dates = sorted(by_date)
        print(f"Parsed range: {dates[0]:%Y-%m-%d} .. {dates[-1]:%Y-%m-%d}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
